"""Deterministic category-aware attribute planning and reference-pack discovery."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

from unilog_product_intelligence.domain.truth import AttributeRecord, ProductTruth

from .models import (
    Applicability,
    AttributePlan,
    AttributeSchema,
    EnrichmentDecision,
    FinalAttributeStatus,
    ReferenceAvailability,
)

EXPECTED_REFERENCE_FILES = (
    "UNILOG_INTERNAL_CONTENT_GUIDELINES.docx",
    "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx",
    "Decimal_Fraction.xlsx",
    "UniCat_Manufacturer_and_Brand_List.xlsx",
    "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx",
    "FAUCETS_LOV.xlsx",
    "Fittings_LOV.xlsx",
    "Reference_Documents_Summary.xlsx",
    "Unilog-Sample_200_Items-Input-vs-Output.xlsx",
    "Sample-1000_Items.xlsx",
)


class ReferencePack:
    """Runtime reference inventory; unavailable data is never replaced with invented values."""

    def __init__(
        self,
        availability: ReferenceAvailability,
        files: dict[str, Path],
        allowed_values: dict[str, tuple[str, ...]] | None = None,
        allowed_uom: dict[str, tuple[str, ...]] | None = None,
    ) -> None:
        self.availability = availability
        self.files = files
        self.allowed_values = allowed_values or {}
        self.allowed_uom = allowed_uom or {}

    @classmethod
    def discover(cls, roots: Iterable[str | Path]) -> ReferencePack:
        found: dict[str, Path] = {}
        for root_value in roots:
            root = Path(root_value)
            if not root.exists():
                continue
            for path in root.rglob("*"):
                if path.is_file() and path.name in EXPECTED_REFERENCE_FILES:
                    found.setdefault(path.name, path)
        allowed_values: dict[str, tuple[str, ...]] = {}
        allowed_uom: dict[str, tuple[str, ...]] = {}
        for filename, path in found.items():
            if path.suffix.casefold() != ".xlsx" or "lov" not in filename.casefold():
                continue
            _load_lov_sheet(path, allowed_values, allowed_uom)
        return cls(
            ReferenceAvailability.REFERENCE_AVAILABLE
            if found
            else ReferenceAvailability.REFERENCE_UNAVAILABLE,
            found,
            allowed_values,
            allowed_uom,
        )

    @property
    def available(self) -> bool:
        return self.availability == ReferenceAvailability.REFERENCE_AVAILABLE


def _load_lov_sheet(
    path: Path, allowed_values: dict[str, tuple[str, ...]], allowed_uom: dict[str, tuple[str, ...]]
) -> None:
    """Best-effort parser for official LOV workbooks without assuming a fixed sheet layout."""

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v).strip().casefold() if v is not None else "" for v in rows[0]]
        attr_index = next(
            (i for i, h in enumerate(headers) if h in {"attribute", "attribute label", "label"}),
            None,
        )
        value_index = next(
            (i for i, h in enumerate(headers) if h in {"value", "allowed value", "lov value"}),
            None,
        )
        uom_index = next((i for i, h in enumerate(headers) if "uom" in h or "unit" in h), None)
        if attr_index is None or value_index is None:
            continue
        values: dict[str, set[str]] = {}
        uoms: dict[str, set[str]] = {}
        for row in rows[1:]:
            if attr_index >= len(row) or value_index >= len(row):
                continue
            attribute = str(row[attr_index]).strip()
            value = str(row[value_index]).strip()
            if not attribute or not value or value.casefold() == "none":
                continue
            values.setdefault(attribute.casefold(), set()).add(value)
            if uom_index is not None and uom_index < len(row) and row[uom_index] not in (None, ""):
                uoms.setdefault(attribute.casefold(), set()).add(str(row[uom_index]).strip())
        for attribute, items in values.items():
            allowed_values[attribute] = tuple(sorted(items))
        for attribute, items in uoms.items():
            allowed_uom[attribute] = tuple(sorted(items))
    workbook.close()


class CategorySchemaRegistry:
    """Small deterministic schema registry; official LOVs can override its constraints."""

    _fitting = (
        AttributeSchema(attribute_id="fitting_type", canonical_name="Fitting Type", required=True),
        AttributeSchema(
            attribute_id="connection_type", canonical_name="Connection Type", required=True
        ),
        AttributeSchema(attribute_id="material", canonical_name="Material", required=True),
        AttributeSchema(
            attribute_id="size", canonical_name="Size", required=True, value_type="measurement"
        ),
    )
    _faucet = (
        AttributeSchema(attribute_id="faucet_type", canonical_name="Faucet Type", required=True),
        AttributeSchema(attribute_id="material", canonical_name="Material", required=False),
        AttributeSchema(attribute_id="finish", canonical_name="Finish", required=False),
        AttributeSchema(
            attribute_id="mounting_type", canonical_name="Mounting Type", required=False
        ),
    )

    def schemas_for(self, product: ProductTruth) -> tuple[AttributeSchema, ...]:
        haystack = " ".join(
            [
                *(product.classification.classpath or ()),
                product.classification.class_name or "",
                product.classification.fine or "",
                str(product.raw_value("Part_Desc") or ""),
            ]
        ).casefold()
        if "faucet" in haystack:
            return self._faucet
        if "fitting" in haystack or "coupling" in haystack or "adapter" in haystack:
            return self._fitting
        return ()


class AttributePlanner:
    """Plans only category-supported, applicable work before any model call."""

    def __init__(
        self,
        reference_pack: ReferencePack | None = None,
        registry: CategorySchemaRegistry | None = None,
    ) -> None:
        self.reference_pack = reference_pack or ReferencePack(
            ReferenceAvailability.REFERENCE_UNAVAILABLE, {}
        )
        self.registry = registry or CategorySchemaRegistry()

    def plan(self, product: ProductTruth) -> tuple[AttributePlan, ...]:
        schemas = self.registry.schemas_for(product)
        existing = {item.attribute_id: item for item in product.attributes}
        if not schemas:
            schemas = tuple(
                AttributeSchema(
                    attribute_id=item.attribute_id,
                    canonical_name=item.canonical_name,
                    required=False,
                    reason="existing ProductTruth attribute; category schema unavailable",
                )
                for item in product.attributes
            )
        plans: list[AttributePlan] = []
        for schema in schemas:
            current = existing.get(schema.attribute_id)
            current_status = _status(current)
            current_value = current.normalized_value if current else None
            evidence_available = bool(current and current.evidence_ids)
            values = self.reference_pack.allowed_values.get(
                schema.canonical_name.casefold(), schema.allowed_values
            )
            uoms = self.reference_pack.allowed_uom.get(
                schema.canonical_name.casefold(), schema.allowed_uom
            )
            if current and current_status in {
                FinalAttributeStatus.VERIFIED,
                FinalAttributeStatus.NORMALIZED,
                FinalAttributeStatus.ENRICHED,
            }:
                decision = EnrichmentDecision.NO_ACTION
                reason = "Existing publishable value is not re-enriched."
            elif evidence_available:
                decision = EnrichmentDecision.VERIFY_EXISTING
                reason = "Existing value has evidence and requires deterministic verification."
            elif schema.required:
                decision = EnrichmentDecision.ENRICH
                reason = "Required category attribute is missing and must be evidence-backed."
            else:
                decision = EnrichmentDecision.ENRICH
                reason = "Optional category attribute may be enriched only when evidence exists."
            plans.append(
                AttributePlan(
                    product_id=product.product_id,
                    category=product.classification.class_name,
                    classpath=product.classification.classpath,
                    attribute_id=schema.attribute_id,
                    attribute_name=schema.canonical_name,
                    applicability=Applicability.REQUIRED
                    if schema.required
                    else Applicability.OPTIONAL,
                    current_status=current_status,
                    current_value=current_value,
                    evidence_available=evidence_available,
                    enrichment_required=decision,
                    validation_requirements=("evidence", "format", "uom")
                    + (("lov",) if values else ()),
                    allowed_values=values,
                    allowed_uom=uoms,
                    reference_availability=self.reference_pack.availability,
                    priority=90 if schema.required else 50,
                    reason=reason,
                )
            )
        return tuple(sorted(plans, key=lambda item: (-item.priority, item.attribute_id)))


def _status(attribute: AttributeRecord | None) -> FinalAttributeStatus:
    if attribute is None or attribute.normalized_value is None:
        return FinalAttributeStatus.MISSING
    mapping = {
        "verified": FinalAttributeStatus.VERIFIED,
        "normalized": FinalAttributeStatus.NORMALIZED,
        "enriched": FinalAttributeStatus.ENRICHED,
        "inferred": FinalAttributeStatus.INFERRED,
        "conflicted": FinalAttributeStatus.CONFLICTED,
        "rejected": FinalAttributeStatus.REJECTED,
    }
    return mapping.get(attribute.status.value, FinalAttributeStatus.REVIEW_REQUIRED)

"""Typed reference-pack loading, normalization, and category-aware taxonomy models.

This module provides dedicated, fail-closed loaders for official UniHack reference files:
- UOM standards and abbreviations
- Decimal to fraction mapping tables
- Manufacturer and brand canonical masters
- Global LOV with hierarchical taxonomy / classpath context
- Category-specific LOV packs (e.g. Faucets, Fittings)
"""

from __future__ import annotations

import re
from collections import defaultdict
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ReferenceAvailability


class ReferenceType(StrEnum):
    UOM_STANDARD = "UOM_STANDARD"
    DECIMAL_FRACTION = "DECIMAL_FRACTION"
    MANUFACTURER_BRAND = "MANUFACTURER_BRAND"
    GLOBAL_LOV = "GLOBAL_LOV"
    CATEGORY_LOV = "CATEGORY_LOV"
    CONTENT_GUIDELINES = "CONTENT_GUIDELINES"
    REFERENCE_DOCUMENTS_SUMMARY = "REFERENCE_DOCUMENTS_SUMMARY"
    DELIVERY_GROUND_TRUTH = "DELIVERY_GROUND_TRUTH"
    SAMPLE_ITEMS = "SAMPLE_ITEMS"


OFFICIAL_REFERENCE_MANIFEST: dict[str, ReferenceType] = {
    "UNILOG_INTERNAL_CONTENT_GUIDELINES.docx": ReferenceType.CONTENT_GUIDELINES,
    "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx": ReferenceType.UOM_STANDARD,
    "Decimal_Fraction.xlsx": ReferenceType.DECIMAL_FRACTION,
    "UniCat_Manufacturer_and_Brand_List.xlsx": ReferenceType.MANUFACTURER_BRAND,
    "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx": ReferenceType.GLOBAL_LOV,
    "FAUCETS_LOV.xlsx": ReferenceType.CATEGORY_LOV,
    "Fittings_LOV.xlsx": ReferenceType.CATEGORY_LOV,
    "Reference_Documents_Summary.xlsx": ReferenceType.REFERENCE_DOCUMENTS_SUMMARY,
    "Unilog-Sample_200_Items-Input-vs-Output.xlsx": ReferenceType.DELIVERY_GROUND_TRUTH,
    "Sample-1000_Items.xlsx": ReferenceType.SAMPLE_ITEMS,
}

EXPECTED_REFERENCE_FILES: tuple[str, ...] = tuple(OFFICIAL_REFERENCE_MANIFEST.keys())


# ==============================================================================
# 1. UOM STANDARDS MODELS & LOADER
# ==============================================================================


@dataclass(frozen=True)
class UomRecord:
    approved_uom: str
    measurement_type: str | None = None
    capture_form: str | None = None
    example: str | None = None
    synonyms: tuple[str, ...] = ()
    source_rule: str | None = None


@dataclass(frozen=True)
class UomStandardMap:
    records: tuple[UomRecord, ...] = ()
    approved_uoms: frozenset[str] = field(default_factory=frozenset)
    canonical_uom_map: dict[str, str] = field(default_factory=dict)
    uoms_by_measurement_type: dict[str, tuple[str, ...]] = field(default_factory=dict)

    def is_approved(self, uom: str) -> bool:
        if not uom:
            return False
        return uom.strip().casefold() in self.approved_uoms

    def normalize(self, uom: str) -> str | None:
        if not uom:
            return None
        clean = uom.strip()
        return self.canonical_uom_map.get(clean.casefold())

    def get_uoms_for_measurement(self, measurement_type: str) -> tuple[str, ...]:
        if not measurement_type:
            return ()
        return self.uoms_by_measurement_type.get(measurement_type.strip().casefold(), ())


def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_uom_master(path: Path) -> UomStandardMap:
    """Load the official UOM Master workbook with dynamic header detection."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return UomStandardMap()

    records: list[UomRecord] = []
    canonical_map: dict[str, str] = {}
    approved_set: set[str] = set()
    by_type: dict[str, set[str]] = defaultdict(set)

    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            # Detect header row
            header_idx = -1
            header_map: dict[str, int] = {}
            for r_idx, row in enumerate(rows[:10]):
                cell_strs = [_clean_str(c).casefold() for c in row]
                if any("approved" in c or "uom" in c or "abbreviation" in c for c in cell_strs):
                    header_idx = r_idx
                    for c_idx, name in enumerate(cell_strs):
                        if not name:
                            continue
                        if (
                            "approved" in name
                            or "standard uom" in name
                            or name in {
                                "uom",
                                "approved uom",
                                "abbreviation",
                                "approved abbreviation",
                            }
                        ):
                            header_map.setdefault("approved_uom", c_idx)
                        elif "measurement" in name or "type" in name:
                            header_map.setdefault("measurement_type", c_idx)
                        elif "capture" in name or "form" in name:
                            header_map.setdefault("capture_form", c_idx)
                        elif "example" in name:
                            header_map.setdefault("example", c_idx)
                        elif "synonym" in name or "alternate" in name or "other" in name:
                            header_map.setdefault("synonyms", c_idx)
                        elif "rule" in name or "term" in name or "description" in name:
                            header_map.setdefault("rule", c_idx)
                    if "approved_uom" in header_map:
                        break

            if header_idx == -1 or "approved_uom" not in header_map:
                continue

            app_col = header_map["approved_uom"]
            meas_col = header_map.get("measurement_type")
            cap_col = header_map.get("capture_form")
            ex_col = header_map.get("example")
            syn_col = header_map.get("synonyms")
            rule_col = header_map.get("rule")

            for row in rows[header_idx + 1 :]:
                if app_col >= len(row):
                    continue
                raw_approved = _clean_str(row[app_col])
                if not raw_approved or raw_approved.casefold() in {"none", "n/a", "null"}:
                    continue

                meas_type = (
                    _clean_str(row[meas_col])
                    if meas_col is not None and meas_col < len(row)
                    else None
                )
                cap_form = (
                    _clean_str(row[cap_col])
                    if cap_col is not None and cap_col < len(row)
                    else None
                )
                ex_val = (
                    _clean_str(row[ex_col])
                    if ex_col is not None and ex_col < len(row)
                    else None
                )
                rule_val = (
                    _clean_str(row[rule_col])
                    if rule_col is not None and rule_col < len(row)
                    else None
                )

                synonyms_list: list[str] = []
                if syn_col is not None and syn_col < len(row):
                    raw_syn = _clean_str(row[syn_col])
                    if raw_syn:
                        for part in re.split(r"[,;/|]", raw_syn):
                            part_clean = part.strip()
                            if part_clean:
                                synonyms_list.append(part_clean)

                record = UomRecord(
                    approved_uom=raw_approved,
                    measurement_type=meas_type or None,
                    capture_form=cap_form or None,
                    example=ex_val or None,
                    synonyms=tuple(synonyms_list),
                    source_rule=rule_val or None,
                )
                records.append(record)
                approved_set.add(raw_approved.casefold())
                canonical_map[raw_approved.casefold()] = raw_approved

                if cap_form:
                    canonical_map[cap_form.casefold()] = raw_approved
                for syn in synonyms_list:
                    canonical_map[syn.casefold()] = raw_approved

                if meas_type:
                    by_type[meas_type.casefold()].add(raw_approved)

    finally:
        wb.close()

    return UomStandardMap(
        records=tuple(records),
        approved_uoms=frozenset(approved_set),
        canonical_uom_map=canonical_map,
        uoms_by_measurement_type={k: tuple(sorted(v)) for k, v in by_type.items()},
    )


# ==============================================================================
# 2. DECIMAL / FRACTION MODELS & LOADER
# ==============================================================================


@dataclass(frozen=True)
class FractionDecimalMap:
    fraction_to_decimal: dict[str, float] = field(default_factory=dict)
    decimal_to_fraction: dict[float, str] = field(default_factory=dict)
    fraction_to_decimal_str: dict[str, str] = field(default_factory=dict)
    decimal_str_to_fraction: dict[str, str] = field(default_factory=dict)

    def to_decimal(self, fraction: str) -> float | None:
        if not fraction:
            return None
        clean = fraction.strip()
        if clean in self.fraction_to_decimal:
            return self.fraction_to_decimal[clean]
        # Try normalized spacing, e.g. "1 1/2" -> "1-1/2"
        alt = clean.replace(" ", "-")
        return self.fraction_to_decimal.get(alt)

    def to_fraction(self, decimal_val: float | str | int | None) -> str | None:
        if decimal_val is None:
            return None
        if isinstance(decimal_val, (int, float)):
            val_float = float(decimal_val)
            if val_float in self.decimal_to_fraction:
                return self.decimal_to_fraction[val_float]
            # Try rounded key
            val_round = round(val_float, 4)
            return self.decimal_to_fraction.get(val_round)
        val_str = str(decimal_val).strip()
        try:
            val_float = float(val_str)
            if val_float in self.decimal_to_fraction:
                return self.decimal_to_fraction[val_float]
        except ValueError:
            pass
        return self.decimal_str_to_fraction.get(val_str)


def load_decimal_fraction(path: Path) -> FractionDecimalMap:
    """Load Decimal_Fraction workbook supporting multiple side-by-side table blocks."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return FractionDecimalMap()

    frac_to_dec: dict[str, float] = {}
    dec_to_frac: dict[float, str] = {}
    frac_to_dec_str: dict[str, str] = {}
    dec_str_to_frac: dict[str, str] = {}

    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            # Find all (fraction_col, decimal_col) pairs in header rows
            pairs: list[tuple[int, int]] = []
            header_row_idx = 0

            for r_idx, row in enumerate(rows[:10]):
                row_cells = [_clean_str(c).casefold() for c in row]
                found_in_row: list[tuple[int, int]] = []
                # Scan adjacent or nearby column pairs
                for c_idx in range(len(row_cells) - 1):
                    c1 = row_cells[c_idx]
                    c2 = row_cells[c_idx + 1]
                    if "fraction" in c1 and "decimal" in c2:
                        found_in_row.append((c_idx, c_idx + 1))
                    elif "decimal" in c1 and "fraction" in c2:
                        found_in_row.append((c_idx + 1, c_idx))
                if found_in_row:
                    pairs = found_in_row
                    header_row_idx = r_idx
                    break

            # If no explicit header was found, infer columns that look like fractions & decimals
            if not pairs:
                for c_idx in range(len(rows[0]) - 1 if rows else 0):
                    pairs.append((c_idx, c_idx + 1))
                header_row_idx = 0

            for row in rows[header_row_idx + 1 :]:
                for frac_col, dec_col in pairs:
                    if frac_col >= len(row) or dec_col >= len(row):
                        continue
                    raw_frac = _clean_str(row[frac_col])
                    raw_dec = _clean_str(row[dec_col])
                    if not raw_frac or not raw_dec:
                        continue
                    if "fraction" in raw_frac.casefold() or "decimal" in raw_dec.casefold():
                        continue

                    try:
                        dec_val = float(raw_dec)
                    except ValueError:
                        continue

                    frac_clean = raw_frac
                    dec_str_clean = raw_dec

                    frac_to_dec[frac_clean] = dec_val
                    dec_to_frac[dec_val] = frac_clean
                    dec_to_frac[round(dec_val, 4)] = frac_clean

                    frac_to_dec_str[frac_clean] = dec_str_clean
                    dec_str_to_frac[dec_str_clean] = frac_clean

                    # Normalized fraction variations, e.g. "1 1/2" and "1-1/2"
                    if " " in frac_clean:
                        alt_frac = frac_clean.replace(" ", "-")
                        frac_to_dec[alt_frac] = dec_val
                        frac_to_dec_str[alt_frac] = dec_str_clean
                    elif "-" in frac_clean:
                        alt_frac = frac_clean.replace("-", " ")
                        frac_to_dec[alt_frac] = dec_val
                        frac_to_dec_str[alt_frac] = dec_str_clean
    finally:
        wb.close()

    return FractionDecimalMap(
        fraction_to_decimal=frac_to_dec,
        decimal_to_fraction=dec_to_frac,
        fraction_to_decimal_str=frac_to_dec_str,
        decimal_str_to_fraction=dec_str_to_frac,
    )


# ==============================================================================
# 3. MANUFACTURER & BRAND MASTER MODELS & LOADER
# ==============================================================================


@dataclass(frozen=True)
class ManufacturerBrandRecord:
    manufacturer_name: str
    manufacturer_code: str | None = None
    brand_name: str | None = None
    brand_code: str | None = None


@dataclass(frozen=True)
class ManufacturerBrandIndex:
    records: tuple[ManufacturerBrandRecord, ...] = ()
    manufacturers_by_name: dict[str, str] = field(default_factory=dict)
    manufacturers_by_code: dict[str, str] = field(default_factory=dict)
    brands_by_name: dict[str, str] = field(default_factory=dict)
    brands_by_code: dict[str, str] = field(default_factory=dict)
    brand_to_manufacturers: dict[str, tuple[str, ...]] = field(default_factory=dict)

    def resolve_manufacturer(self, query: str) -> str | None:
        if not query:
            return None
        clean = query.strip()
        key = clean.casefold()
        if key in self.manufacturers_by_name:
            return self.manufacturers_by_name[key]
        if key in self.manufacturers_by_code:
            return self.manufacturers_by_code[key]
        # Strip common distributor/legal suffixes for lookup
        stripped = re.sub(r"[,.]?\s+(?:llc|inc|corp|corporation|co|ltd)\b.*", "", key).strip()
        if stripped in self.manufacturers_by_name:
            return self.manufacturers_by_name[stripped]
        return None

    def resolve_brand(self, query: str) -> str | None:
        if not query:
            return None
        clean = query.strip()
        key = clean.casefold()
        if key in self.brands_by_name:
            return self.brands_by_name[key]
        if key in self.brands_by_code:
            return self.brands_by_code[key]
        stripped = re.sub(r"[™®©]", "", key).strip()
        if stripped in self.brands_by_name:
            return self.brands_by_name[stripped]
        return None

    def resolve_pair(
        self, brand: str | None, manufacturer: str | None = None
    ) -> tuple[str | None, str | None]:
        res_mfg = self.resolve_manufacturer(manufacturer) if manufacturer else None
        res_brand = self.resolve_brand(brand) if brand else None

        if res_brand and not res_mfg:
            mfgs = self.brand_to_manufacturers.get(res_brand.casefold(), ())
            if len(mfgs) == 1:
                res_mfg = mfgs[0]
        return res_mfg, res_brand


def load_manufacturer_brand(path: Path) -> ManufacturerBrandIndex:
    """Load UniCat Manufacturer and Brand list preserving exact canonical casing."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ManufacturerBrandIndex()

    records: list[ManufacturerBrandRecord] = []
    mfgs_by_name: dict[str, str] = {}
    mfgs_by_code: dict[str, str] = {}
    brands_by_name: dict[str, str] = {}
    brands_by_code: dict[str, str] = {}
    b_to_m: dict[str, set[str]] = defaultdict(set)

    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = -1
            col_map: dict[str, int] = {}
            for r_idx, row in enumerate(rows[:10]):
                cell_strs = [_clean_str(c).casefold() for c in row]
                if any("manufacturer" in c or "brand" in c for c in cell_strs):
                    header_idx = r_idx
                    for c_idx, name in enumerate(cell_strs):
                        if not name:
                            continue
                        if "manufacturer" in name and "code" in name:
                            col_map.setdefault("mfg_code", c_idx)
                        elif "manufacturer" in name and ("name" in name or name == "manufacturer"):
                            col_map.setdefault("mfg_name", c_idx)
                        elif "brand" in name and "code" in name:
                            col_map.setdefault("brand_code", c_idx)
                        elif "brand" in name and ("name" in name or name == "brand"):
                            col_map.setdefault("brand_name", c_idx)
                    if "mfg_name" in col_map or "brand_name" in col_map:
                        break

            if header_idx == -1:
                continue

            mfg_name_col = col_map.get("mfg_name")
            mfg_code_col = col_map.get("mfg_code")
            brand_name_col = col_map.get("brand_name")
            brand_code_col = col_map.get("brand_code")

            for row in rows[header_idx + 1 :]:
                mfg_name = (
                    _clean_str(row[mfg_name_col])
                    if mfg_name_col is not None and mfg_name_col < len(row)
                    else ""
                )
                mfg_code = (
                    _clean_str(row[mfg_code_col])
                    if mfg_code_col is not None and mfg_code_col < len(row)
                    else ""
                )
                brand_name = (
                    _clean_str(row[brand_name_col])
                    if brand_name_col is not None and brand_name_col < len(row)
                    else ""
                )
                brand_code = (
                    _clean_str(row[brand_code_col])
                    if brand_code_col is not None and brand_code_col < len(row)
                    else ""
                )

                if not mfg_name and not brand_name:
                    continue

                rec = ManufacturerBrandRecord(
                    manufacturer_name=mfg_name,
                    manufacturer_code=mfg_code or None,
                    brand_name=brand_name or None,
                    brand_code=brand_code or None,
                )
                records.append(rec)

                if mfg_name:
                    mfgs_by_name[mfg_name.casefold()] = mfg_name
                    base_mfg = re.sub(
                        r"[,.]?\s+(?:llc|inc|corp|corporation|co|ltd)\b.*",
                        "",
                        mfg_name.casefold(),
                    ).strip()
                    if base_mfg and base_mfg != mfg_name.casefold():
                        mfgs_by_name.setdefault(base_mfg, mfg_name)
                if mfg_code:
                    mfgs_by_code[mfg_code.casefold()] = mfg_name or mfg_code
                if brand_name:
                    brands_by_name[brand_name.casefold()] = brand_name
                    clean_brand = re.sub(r"[™®©]", "", brand_name.casefold()).strip()
                    if clean_brand and clean_brand != brand_name.casefold():
                        brands_by_name.setdefault(clean_brand, brand_name)
                    if mfg_name:
                        b_to_m[brand_name.casefold()].add(mfg_name)
                        if clean_brand:
                            b_to_m[clean_brand].add(mfg_name)
                if brand_code:
                    brands_by_code[brand_code.casefold()] = brand_name or brand_code
    finally:
        wb.close()

    return ManufacturerBrandIndex(
        records=tuple(records),
        manufacturers_by_name=mfgs_by_name,
        manufacturers_by_code=mfgs_by_code,
        brands_by_name=brands_by_name,
        brands_by_code=brands_by_code,
        brand_to_manufacturers={k: tuple(sorted(v)) for k, v in b_to_m.items()},
    )


# ==============================================================================
# 4. GLOBAL & CATEGORY LOV MODELS & LOADERS
# ==============================================================================


@dataclass(frozen=True)
class LovAttributeRule:
    classpath: tuple[str, ...]
    leaf_node: str | None
    attribute_label: str
    attribute_values: tuple[str, ...]
    normalized_label: str | None = None
    normalized_values: tuple[str, ...] = ()
    filtering: bool | None = None
    guidelines: str | None = None
    remarks: str | None = None
    allowed_uom: tuple[str, ...] = ()


def _normalize_classpath_str(cp: str | tuple[str, ...] | None) -> str:
    if not cp:
        return ""
    if isinstance(cp, tuple):
        return " > ".join(s.strip() for s in cp if s.strip()).casefold()
    return cp.strip().casefold()


@dataclass(frozen=True)
class GlobalLovIndex:
    rules: tuple[LovAttributeRule, ...] = ()
    rules_by_classpath_attr: dict[tuple[str, str], LovAttributeRule] = field(default_factory=dict)
    rules_by_leaf_attr: dict[tuple[str, str], LovAttributeRule] = field(default_factory=dict)
    rules_by_attr: dict[str, list[LovAttributeRule]] = field(default_factory=dict)
    rules_by_classpath: dict[str, list[LovAttributeRule]] = field(default_factory=dict)
    rules_by_leaf: dict[str, list[LovAttributeRule]] = field(default_factory=dict)

    def get_rule(
        self,
        attribute: str,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> LovAttributeRule | None:
        if not attribute:
            return None
        attr_key = attribute.strip().casefold()

        # 1. Exact classpath match
        if classpath:
            cp_key = _normalize_classpath_str(classpath)
            rule = self.rules_by_classpath_attr.get((cp_key, attr_key))
            if rule:
                return rule

        # 2. Category / Leaf Node match
        if category:
            cat_key = category.strip().casefold()
            rule = self.rules_by_leaf_attr.get((cat_key, attr_key))
            if rule:
                return rule

        # 3. If only one rule exists across all classpaths for this attribute, return it
        all_for_attr = self.rules_by_attr.get(attr_key, [])
        if len(all_for_attr) == 1:
            return all_for_attr[0]

        return None

    def get_allowed_values(
        self,
        attribute: str,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> tuple[str, ...]:
        rule = self.get_rule(attribute, classpath, category)
        if rule:
            return rule.attribute_values
        # If no specific rule matched but attribute exists across multiple
        # categories, do NOT conflate
        return ()

    def get_allowed_uom(
        self,
        attribute: str,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> tuple[str, ...]:
        rule = self.get_rule(attribute, classpath, category)
        if rule:
            return rule.allowed_uom
        return ()

    def get_rules_for_category(
        self, category_or_classpath: str | tuple[str, ...]
    ) -> tuple[LovAttributeRule, ...]:
        if not category_or_classpath:
            return ()
        if isinstance(category_or_classpath, tuple):
            cp_key = _normalize_classpath_str(category_or_classpath)
            if cp_key in self.rules_by_classpath:
                return tuple(self.rules_by_classpath[cp_key])
            return tuple(
                rule
                for (cp, _), rule in self.rules_by_classpath_attr.items()
                if cp == cp_key or cp_key in cp or cp in cp_key
            )
        cat_key = category_or_classpath.strip().casefold()
        if cat_key in self.rules_by_leaf:
            return tuple(self.rules_by_leaf[cat_key])
        return tuple(
            rule for (leaf, _), rule in self.rules_by_leaf_attr.items() if leaf == cat_key
        )


@dataclass(frozen=True)
class CategoryLovPack:
    category_name: str
    attribute_rules: dict[str, LovAttributeRule] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)

    def get_allowed_values(self, attribute: str) -> tuple[str, ...]:
        rule = self.attribute_rules.get(attribute.strip().casefold())
        return rule.attribute_values if rule else ()

    def get_allowed_uom(self, attribute: str) -> tuple[str, ...]:
        rule = self.attribute_rules.get(attribute.strip().casefold())
        return rule.allowed_uom if rule else ()


def _parse_lov_rows(rows: Sequence[Sequence[Any]]) -> list[LovAttributeRule]:
    """Parse table rows into structured LovAttributeRule entries."""
    if not rows:
        return []

    # Detect header row
    header_idx = -1
    col_map: dict[str, int] = {}

    for r_idx, row in enumerate(rows[:10]):
        cell_strs = [_clean_str(c).casefold() for c in row]
        if any("attribute" in c or "classpath" in c or "label" in c for c in cell_strs):
            header_idx = r_idx
            for c_idx, name in enumerate(cell_strs):
                if not name:
                    continue
                if "classpath" in name or "class path" in name or "taxonomy" in name:
                    col_map.setdefault("classpath", c_idx)
                elif "leaf" in name or "leaf node" in name or "category" in name:
                    col_map.setdefault("leaf_node", c_idx)
                elif "filter" in name:
                    col_map.setdefault("filtering", c_idx)
                elif "attribute values" in name or "allowed values" in name or "lov values" in name:
                    col_map.setdefault("values", c_idx)
                elif "attribute" in name or "attribute label" in name or name in {"label", "attr"}:
                    col_map.setdefault("attribute", c_idx)
                elif "normalized label" in name or "canonical label" in name:
                    col_map.setdefault("norm_label", c_idx)
                elif "normalized value" in name or "normalized values" in name:
                    col_map.setdefault("norm_values", c_idx)
                elif "guideline" in name or "guidelines" in name:
                    col_map.setdefault("guidelines", c_idx)
                elif "remark" in name or "remarks" in name or "comment" in name or "note" in name:
                    col_map.setdefault("remarks", c_idx)
                elif "uom" in name or "unit" in name:
                    col_map.setdefault("uom", c_idx)
            if "attribute" in col_map:
                break

    if header_idx == -1 or "attribute" not in col_map:
        return []

    attr_col = col_map["attribute"]
    cp_col = col_map.get("classpath")
    leaf_col = col_map.get("leaf_node")
    filt_col = col_map.get("filtering")
    val_col = col_map.get("values")
    nlabel_col = col_map.get("norm_label")
    nval_col = col_map.get("norm_values")
    guide_col = col_map.get("guidelines")
    rem_col = col_map.get("remarks")
    uom_col = col_map.get("uom")

    # Group values by (classpath, leaf_node, attribute)
    # This handles multi-row spreadsheets where each value is on its own row,
    # and merged cells where classpath or attribute label is forward-filled.
    last_cp: str = ""
    last_leaf: str = ""
    last_attr: str = ""

    grouped_data: dict[
        tuple[str, str, str],
        dict[str, Any],
    ] = {}

    for row in rows[header_idx + 1 :]:
        raw_cp = _clean_str(row[cp_col]) if cp_col is not None and cp_col < len(row) else ""
        raw_leaf = _clean_str(row[leaf_col]) if leaf_col is not None and leaf_col < len(row) else ""
        raw_attr = _clean_str(row[attr_col]) if attr_col < len(row) else ""
        raw_val = _clean_str(row[val_col]) if val_col is not None and val_col < len(row) else ""
        raw_nval = _clean_str(row[nval_col]) if nval_col is not None and nval_col < len(row) else ""
        raw_uom = _clean_str(row[uom_col]) if uom_col is not None and uom_col < len(row) else ""
        raw_filt = _clean_str(row[filt_col]) if filt_col is not None and filt_col < len(row) else ""
        raw_guide = (
            _clean_str(row[guide_col]) if guide_col is not None and guide_col < len(row) else ""
        )
        raw_rem = _clean_str(row[rem_col]) if rem_col is not None and rem_col < len(row) else ""
        raw_nlabel = (
            _clean_str(row[nlabel_col]) if nlabel_col is not None and nlabel_col < len(row) else ""
        )

        # Merged / forward-fill logic: if attribute is empty but value exists, use last attribute
        if raw_cp:
            last_cp = raw_cp
        if raw_leaf:
            last_leaf = raw_leaf
        if raw_attr:
            last_attr = raw_attr

        current_cp = raw_cp or last_cp
        current_leaf = raw_leaf or last_leaf
        current_attr = raw_attr or last_attr

        if not current_attr:
            continue

        is_filt = (
            True if raw_filt.casefold() in {"y", "yes", "true", "1"}
            else False if raw_filt.casefold() in {"n", "no", "false", "0"}
            else None
        )

        group_key = (current_cp, current_leaf, current_attr)
        if group_key not in grouped_data:
            grouped_data[group_key] = {
                "classpath": tuple(
                    s.strip() for s in re.split(r"[>/|]", current_cp) if s.strip()
                ),
                "leaf_node": current_leaf or None,
                "attribute_label": current_attr,
                "values": set(),
                "norm_values": set(),
                "uoms": set(),
                "normalized_label": raw_nlabel or None,
                "filtering": is_filt,
                "guidelines": raw_guide or None,
                "remarks": raw_rem or None,
            }

        entry = grouped_data[group_key]
        if raw_val and raw_val.casefold() not in {"none", "n/a", "null"}:
            for sub_val in re.split(r"[\n\r]", raw_val):
                sub_clean = sub_val.strip()
                if sub_clean:
                    entry["values"].add(sub_clean)

        if raw_nval and raw_nval.casefold() not in {"none", "n/a", "null"}:
            for sub_nval in re.split(r"[\n\r]", raw_nval):
                sub_nclean = sub_nval.strip()
                if sub_nclean:
                    entry["norm_values"].add(sub_nclean)

        if raw_uom and raw_uom.casefold() not in {"none", "n/a", "null"}:
            for sub_uom in re.split(r"[,;/\n\r]", raw_uom):
                sub_uom_clean = sub_uom.strip()
                if sub_uom_clean:
                    entry["uoms"].add(sub_uom_clean)

    rules: list[LovAttributeRule] = []
    for entry in grouped_data.values():
        rules.append(
            LovAttributeRule(
                classpath=entry["classpath"],
                leaf_node=entry["leaf_node"],
                attribute_label=entry["attribute_label"],
                attribute_values=tuple(sorted(entry["values"])),
                normalized_label=entry["normalized_label"],
                normalized_values=tuple(sorted(entry["norm_values"])),
                filtering=entry["filtering"],
                guidelines=entry["guidelines"],
                remarks=entry["remarks"],
                allowed_uom=tuple(sorted(entry["uoms"])),
            )
        )
    return rules


def load_global_lov(path: Path) -> GlobalLovIndex:
    """Load the official Global LOV workbook with hierarchical classpath context."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return GlobalLovIndex()

    all_rules: list[LovAttributeRule] = []
    by_cp_attr: dict[tuple[str, str], LovAttributeRule] = {}
    by_leaf_attr: dict[tuple[str, str], LovAttributeRule] = {}
    by_attr: dict[str, list[LovAttributeRule]] = defaultdict(list)
    by_cp: dict[str, list[LovAttributeRule]] = defaultdict(list)
    by_leaf: dict[str, list[LovAttributeRule]] = defaultdict(list)

    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            sheet_rules = _parse_lov_rows(rows)
            for rule in sheet_rules:
                all_rules.append(rule)
                attr_key = rule.attribute_label.casefold()
                if rule.classpath:
                    cp_key = _normalize_classpath_str(rule.classpath)
                    by_cp_attr[(cp_key, attr_key)] = rule
                    by_cp[cp_key].append(rule)
                if rule.leaf_node:
                    leaf_key = rule.leaf_node.casefold()
                    by_leaf_attr[(leaf_key, attr_key)] = rule
                    by_leaf[leaf_key].append(rule)
                by_attr[attr_key].append(rule)
    finally:
        wb.close()

    return GlobalLovIndex(
        rules=tuple(all_rules),
        rules_by_classpath_attr=by_cp_attr,
        rules_by_leaf_attr=by_leaf_attr,
        rules_by_attr=dict(by_attr),
        rules_by_classpath=dict(by_cp),
        rules_by_leaf=dict(by_leaf),
    )


def load_category_lov(path: Path, category_name: str | None = None) -> CategoryLovPack:
    """Load category-specific LOV workbook (e.g. Faucets, Fittings)."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return CategoryLovPack(category_name=category_name or path.stem)

    rules_dict: dict[str, LovAttributeRule] = {}
    meta: dict[str, Any] = {"sheets": wb.sheetnames}
    cat_name = category_name or path.stem.replace("_LOV", "").replace("_lov", "").title()

    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            sheet_rules = _parse_lov_rows(rows)
            for rule in sheet_rules:
                rules_dict[rule.attribute_label.casefold()] = rule
    finally:
        wb.close()

    return CategoryLovPack(
        category_name=cat_name,
        attribute_rules=rules_dict,
        metadata=meta,
    )


# ==============================================================================
# 5. REFERENCE PACK AGGREGATE CONTAINER
# ==============================================================================


class ReferencePack:
    """Runtime reference inventory with per-reference availability and taxonomy awareness."""

    def __init__(
        self,
        availability: ReferenceAvailability = ReferenceAvailability.REFERENCE_UNAVAILABLE,
        files: dict[str, Path] | None = None,
        allowed_values: dict[str, tuple[str, ...]] | None = None,
        allowed_uom: dict[str, tuple[str, ...]] | None = None,
        uom_standards: UomStandardMap | None = None,
        decimal_fractions: FractionDecimalMap | None = None,
        manufacturer_brands: ManufacturerBrandIndex | None = None,
        global_lov: GlobalLovIndex | None = None,
        category_lovs: dict[str, CategoryLovPack] | None = None,
        status: dict[ReferenceType, ReferenceAvailability] | None = None,
    ) -> None:
        self.files = files or {}
        self.uom_standards = uom_standards
        self.decimal_fractions = decimal_fractions
        self.manufacturer_brands = manufacturer_brands
        self.global_lov = global_lov
        self.category_lovs = category_lovs or {}
        self.status = status or {
            t: (
                ReferenceAvailability.REFERENCE_AVAILABLE
                if self._is_type_present(t)
                else ReferenceAvailability.REFERENCE_UNAVAILABLE
            )
            for t in ReferenceType
        }
        # Overall availability is AVAILABLE if at least one reference is present
        self.availability = availability

        # Aggregate fallback dictionaries for legacy / backwards compatibility
        self._allowed_values = allowed_values
        self._allowed_uom = allowed_uom

    def _is_type_present(self, ref_type: ReferenceType) -> bool:
        if ref_type == ReferenceType.UOM_STANDARD:
            return bool(self.uom_standards and self.uom_standards.records)
        if ref_type == ReferenceType.DECIMAL_FRACTION:
            return bool(self.decimal_fractions and self.decimal_fractions.fraction_to_decimal)
        if ref_type == ReferenceType.MANUFACTURER_BRAND:
            return bool(self.manufacturer_brands and self.manufacturer_brands.records)
        if ref_type == ReferenceType.GLOBAL_LOV:
            return bool(self.global_lov and self.global_lov.rules)
        if ref_type == ReferenceType.CATEGORY_LOV:
            return bool(self.category_lovs)
        return any(
            OFFICIAL_REFERENCE_MANIFEST.get(fname) == ref_type
            for fname in self.files
        )

    @classmethod
    def discover(cls, roots: Iterable[str | Path]) -> ReferencePack:
        found: dict[str, Path] = {}
        for root_value in roots:
            root = Path(root_value)
            if not root.exists():
                continue
            for path in root.rglob("*"):
                if not path.is_file() or path.name not in OFFICIAL_REFERENCE_MANIFEST:
                    continue
                try:
                    rel_parts = path.relative_to(root).parts
                except ValueError:
                    rel_parts = path.parts
                if not any(part.startswith(".") for part in rel_parts[:-1]):
                    found.setdefault(path.name, path)

        uom_map: UomStandardMap | None = None
        dec_frac_map: FractionDecimalMap | None = None
        mfg_brand_index: ManufacturerBrandIndex | None = None
        global_lov_index: GlobalLovIndex | None = None
        category_lovs: dict[str, CategoryLovPack] = {}
        status_map: dict[ReferenceType, ReferenceAvailability] = {
            t: ReferenceAvailability.REFERENCE_UNAVAILABLE for t in ReferenceType
        }

        # Load each discovered file with its dedicated loader
        for filename, path in found.items():
            ref_type = OFFICIAL_REFERENCE_MANIFEST.get(filename)
            if ref_type == ReferenceType.UOM_STANDARD:
                uom_map = load_uom_master(path)
                if uom_map.records:
                    status_map[ReferenceType.UOM_STANDARD] = (
                        ReferenceAvailability.REFERENCE_AVAILABLE
                    )
            elif ref_type == ReferenceType.DECIMAL_FRACTION:
                dec_frac_map = load_decimal_fraction(path)
                if dec_frac_map.fraction_to_decimal:
                    status_map[ReferenceType.DECIMAL_FRACTION] = (
                        ReferenceAvailability.REFERENCE_AVAILABLE
                    )
            elif ref_type == ReferenceType.MANUFACTURER_BRAND:
                mfg_brand_index = load_manufacturer_brand(path)
                if mfg_brand_index.records:
                    status_map[ReferenceType.MANUFACTURER_BRAND] = (
                        ReferenceAvailability.REFERENCE_AVAILABLE
                    )
            elif ref_type == ReferenceType.GLOBAL_LOV:
                global_lov_index = load_global_lov(path)
                if global_lov_index.rules:
                    status_map[ReferenceType.GLOBAL_LOV] = (
                        ReferenceAvailability.REFERENCE_AVAILABLE
                    )
            elif ref_type == ReferenceType.CATEGORY_LOV:
                cat_pack = load_category_lov(path)
                if cat_pack.attribute_rules:
                    category_lovs[cat_pack.category_name.casefold()] = cat_pack
                    status_map[ReferenceType.CATEGORY_LOV] = (
                        ReferenceAvailability.REFERENCE_AVAILABLE
                    )
            else:
                if ref_type:
                    status_map[ref_type] = ReferenceAvailability.REFERENCE_AVAILABLE

        overall_availability = (
            ReferenceAvailability.REFERENCE_AVAILABLE
            if any(st == ReferenceAvailability.REFERENCE_AVAILABLE for st in status_map.values())
            else ReferenceAvailability.REFERENCE_UNAVAILABLE
        )

        return cls(
            availability=overall_availability,
            files=found,
            uom_standards=uom_map,
            decimal_fractions=dec_frac_map,
            manufacturer_brands=mfg_brand_index,
            global_lov=global_lov_index,
            category_lovs=category_lovs,
            status=status_map,
        )

    @property
    def available(self) -> bool:
        return self.availability == ReferenceAvailability.REFERENCE_AVAILABLE

    @property
    def uom_available(self) -> bool:
        return (
            self.status.get(ReferenceType.UOM_STANDARD)
            == ReferenceAvailability.REFERENCE_AVAILABLE
        )

    @property
    def lov_available(self) -> bool:
        return (
            self.status.get(ReferenceType.GLOBAL_LOV)
            == ReferenceAvailability.REFERENCE_AVAILABLE
            or self.status.get(ReferenceType.CATEGORY_LOV)
            == ReferenceAvailability.REFERENCE_AVAILABLE
        )

    @property
    def manufacturer_brand_available(self) -> bool:
        return (
            self.status.get(ReferenceType.MANUFACTURER_BRAND)
            == ReferenceAvailability.REFERENCE_AVAILABLE
        )

    @property
    def decimal_fraction_available(self) -> bool:
        return (
            self.status.get(ReferenceType.DECIMAL_FRACTION)
            == ReferenceAvailability.REFERENCE_AVAILABLE
        )

    @property
    def allowed_values(self) -> dict[str, tuple[str, ...]]:
        """Legacy flat allowed_values dictionary."""
        if self._allowed_values is not None:
            return self._allowed_values
        res: dict[str, set[str]] = defaultdict(set)
        if self.global_lov:
            for rule in self.global_lov.rules:
                res[rule.attribute_label.casefold()].update(rule.attribute_values)
        for cat_pack in self.category_lovs.values():
            for attr, rule in cat_pack.attribute_rules.items():
                res[attr.casefold()].update(rule.attribute_values)
        return {k: tuple(sorted(v)) for k, v in res.items()}

    @property
    def allowed_uom(self) -> dict[str, tuple[str, ...]]:
        """Legacy flat allowed_uom dictionary."""
        if self._allowed_uom is not None:
            return self._allowed_uom
        res: dict[str, set[str]] = defaultdict(set)
        if self.global_lov:
            for rule in self.global_lov.rules:
                if rule.allowed_uom:
                    res[rule.attribute_label.casefold()].update(rule.allowed_uom)
        for cat_pack in self.category_lovs.values():
            for attr, rule in cat_pack.attribute_rules.items():
                if rule.allowed_uom:
                    res[attr.casefold()].update(rule.allowed_uom)
        return {k: tuple(sorted(v)) for k, v in res.items()}

    def get_allowed_values(
        self,
        attribute: str,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> tuple[str, ...]:
        """Category-aware allowed values lookup with fail-closed semantics."""
        if not attribute:
            return ()
        attr_clean = attribute.strip().casefold()

        # 1. Category-specific LOV pack match
        if category:
            cat_pack = self.category_lovs.get(category.strip().casefold())
            if cat_pack:
                vals = cat_pack.get_allowed_values(attr_clean)
                if vals:
                    return vals

        # 2. Global LOV with classpath / category context
        if self.global_lov:
            vals = self.global_lov.get_allowed_values(
                attr_clean, classpath=classpath, category=category
            )
            if vals:
                return vals

        # 3. Fallback to legacy dictionary if populated directly
        if self._allowed_values and attr_clean in self._allowed_values:
            return self._allowed_values[attr_clean]

        return ()

    def get_allowed_uom(
        self,
        attribute: str,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> tuple[str, ...]:
        """Category-aware allowed UOM lookup with fail-closed semantics."""
        if not attribute:
            return ()
        attr_clean = attribute.strip().casefold()

        # 1. Category-specific LOV pack match
        if category:
            cat_pack = self.category_lovs.get(category.strip().casefold())
            if cat_pack:
                uoms = cat_pack.get_allowed_uom(attr_clean)
                if uoms:
                    return uoms

        # 2. Global LOV with classpath / category context
        if self.global_lov:
            uoms = self.global_lov.get_allowed_uom(
                attr_clean, classpath=classpath, category=category
            )
            if uoms:
                return uoms

        # 3. Fallback to legacy dictionary if populated directly
        if self._allowed_uom and attr_clean in self._allowed_uom:
            return self._allowed_uom[attr_clean]

        return ()

    def get_attribute_rule(
        self,
        attribute: str,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> LovAttributeRule | None:
        """Lookup full LovAttributeRule including filtering, guidelines, remarks."""
        if not attribute:
            return None
        attr_clean = attribute.strip().casefold()

        if category:
            cat_pack = self.category_lovs.get(category.strip().casefold())
            if cat_pack and attr_clean in cat_pack.attribute_rules:
                return cat_pack.attribute_rules[attr_clean]

        if self.global_lov:
            return self.global_lov.get_rule(attr_clean, classpath=classpath, category=category)

        return None

    def normalize_uom(self, uom: str) -> str | None:
        if self.uom_standards:
            return self.uom_standards.normalize(uom)
        return None

    def is_approved_uom(self, uom: str) -> bool:
        if self.uom_standards:
            return self.uom_standards.is_approved(uom)
        return False

    def fraction_to_decimal(self, fraction: str) -> float | None:
        if self.decimal_fractions:
            return self.decimal_fractions.to_decimal(fraction)
        return None

    def decimal_to_fraction(self, decimal_val: float | str | int) -> str | None:
        if self.decimal_fractions:
            return self.decimal_fractions.to_fraction(decimal_val)
        return None

    def resolve_manufacturer(self, name_or_code: str) -> str | None:
        if self.manufacturer_brands:
            return self.manufacturer_brands.resolve_manufacturer(name_or_code)
        return None

    def resolve_brand(self, name_or_code: str) -> str | None:
        if self.manufacturer_brands:
            return self.manufacturer_brands.resolve_brand(name_or_code)
        return None

    def resolve_brand_manufacturer_pair(
        self, brand: str | None, manufacturer: str | None = None
    ) -> tuple[str | None, str | None]:
        if self.manufacturer_brands:
            return self.manufacturer_brands.resolve_pair(brand, manufacturer)
        return None, None

    def resolve_category_rules(
        self,
        classpath: tuple[str, ...] | None = None,
        category: str | None = None,
    ) -> tuple[tuple[LovAttributeRule, ...], str] | tuple[None, str]:
        """Resolve attribute rules following the 5-tier resolution hierarchy."""
        # Priority 1: Exact category-specific LOV pack
        if category:
            cat_key = category.strip().casefold()
            cat_pack = self.category_lovs.get(cat_key)
            if cat_pack and cat_pack.attribute_rules:
                return tuple(cat_pack.attribute_rules.values()), "CATEGORY_LOV"
            cat_stem = cat_key.rstrip("s")
            for pack_key, pack_obj in self.category_lovs.items():
                if (
                    pack_key == cat_key or pack_key.rstrip("s") == cat_stem
                ) and pack_obj.attribute_rules:
                    return tuple(pack_obj.attribute_rules.values()), "CATEGORY_LOV"

        if classpath:
            for part in classpath:
                part_key = part.strip().casefold()
                part_pack = self.category_lovs.get(part_key)
                if part_pack and part_pack.attribute_rules:
                    return tuple(part_pack.attribute_rules.values()), "CATEGORY_LOV"

        # Priority 2: Exact classpath + leaf-node rules from GLOBAL_LOV
        if self.global_lov:
            if classpath and category:
                cp_key = _normalize_classpath_str(classpath)
                cat_key = category.strip().casefold()
                cp_rules = self.global_lov.rules_by_classpath.get(cp_key, [])
                exact_rules = [
                    r
                    for r in cp_rules
                    if r.leaf_node and r.leaf_node.strip().casefold() == cat_key
                ]
                if exact_rules:
                    return tuple(exact_rules), "GLOBAL_LOV"

            # Priority 3: Classpath-level rules from GLOBAL_LOV
            if classpath:
                cp_key = _normalize_classpath_str(classpath)
                if cp_key in self.global_lov.rules_by_classpath:
                    rules = self.global_lov.rules_by_classpath[cp_key]
                    if rules:
                        return tuple(rules), "GLOBAL_LOV"
                for stored_cp, rules in self.global_lov.rules_by_classpath.items():
                    if (cp_key in stored_cp or stored_cp in cp_key) and rules:
                        return tuple(rules), "GLOBAL_LOV"

            # Priority 4: Category/leaf-name fallback lookup in GLOBAL_LOV
            if category:
                cat_key = category.strip().casefold()
                if cat_key in self.global_lov.rules_by_leaf:
                    rules = self.global_lov.rules_by_leaf[cat_key]
                    if rules:
                        return tuple(rules), "GLOBAL_LOV"
                for leaf_key, rules in self.global_lov.rules_by_leaf.items():
                    if (cat_key in leaf_key or leaf_key in cat_key) and rules:
                        return tuple(rules), "GLOBAL_LOV"

        # Priority 5: Fallback
        return None, "FALLBACK_EXISTING_ATTRIBUTES"


_MEASUREMENT_UNIT_RE = re.compile(
    r"^([+-]?(?:\d+(?:[./]\d+)?|\d+\s+\d+/\d+))\s*([a-zA-Z°%\"'#].*)$"
)


def separate_value_and_uom(
    raw: Any,
    allowed_uoms: tuple[str, ...] = (),
    reference_pack: ReferencePack | None = None,
) -> tuple[Any, str | None]:
    """Deterministically separate a numeric/fractional value and its trailing UOM."""
    if raw is None or isinstance(raw, (int, float, bool)):
        return raw, None
    raw_str = str(raw).strip()
    if not raw_str:
        return raw, None

    match = _MEASUREMENT_UNIT_RE.match(raw_str)
    if match:
        val_part = match.group(1).strip()
        unit_part = match.group(2).strip()

        if reference_pack:
            norm_uom = reference_pack.normalize_uom(unit_part)
            if norm_uom:
                return val_part, norm_uom

        for allowed in allowed_uoms:
            if allowed and allowed.strip().casefold() == unit_part.casefold():
                return val_part, allowed

        return val_part, unit_part

    return raw, None

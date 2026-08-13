"""Format-specific readers kept separate from business and persistence logic."""

import csv
import hashlib
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .contracts import DatasetRow, FileFormat, SheetInspection, SourceFile
from .normalize import normalize_row


class TabularReadResult:
    """Internal reader result shared by CSV and XLSX adapters."""

    def __init__(
        self, source_file: SourceFile, sheets: list[SheetInspection], rows: list[DatasetRow]
    ):
        self.source_file = source_file
        self.sheets = sheets
        self.rows = rows


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _format_for(path: Path) -> FileFormat:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return FileFormat.CSV
    if suffix == ".xlsx":
        return FileFormat.XLSX
    raise ValueError(f"Unsupported tabular format: {path.suffix}")


def _source_file(path: Path) -> SourceFile:
    return SourceFile(
        path=str(path),
        name=path.name,
        format=_format_for(path),
        available=True,
        size_bytes=path.stat().st_size,
        sha256=_sha256(path),
    )


def _unique_headers(headers: Iterable[Any]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for position, value in enumerate(headers, start=1):
        name = "" if value is None else str(value).strip()
        if name == "":
            name = f"__unnamed_{position}"
        counts[name] = counts.get(name, 0) + 1
        result.append(name if counts[name] == 1 else f"{name}#{counts[name]}")
    return result


def _row_hash(values: list[Any]) -> str:
    raw = "\x1f".join("" if value is None else str(value) for value in values)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _dataset_rows(
    headers: list[str], rows: Iterable[Iterable[Any]], start_row: int
) -> list[DatasetRow]:
    result: list[DatasetRow] = []
    for row_number, row in enumerate(rows, start=start_row):
        values = list(row)
        padded = values + [None] * max(0, len(headers) - len(values))
        raw_values = {
            header: padded[index] if index < len(padded) else None
            for index, header in enumerate(headers)
        }
        normalization = normalize_row(raw_values)
        result.append(
            DatasetRow(
                row_number=row_number,
                raw_values=raw_values,
                normalized_values={
                    key: value.normalized_value for key, value in normalization.items()
                },
                normalization=normalization,
                row_hash=_row_hash(values),
            )
        )
    return result


def _csv_read(path: Path) -> TabularReadResult:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    headers = _unique_headers(rows[0] if rows else [])
    data_rows = rows[1:] if rows else []
    dataset_rows = _dataset_rows(headers, data_rows, start_row=2)
    sheet = SheetInspection(
        name=path.stem,
        header_row_index=1,
        headers=headers,
        row_count=len(data_rows),
        column_count=len(headers),
        leading_rows=rows[:5],
    )
    return TabularReadResult(_source_file(path), [sheet], dataset_rows)


def _header_row(values: list[tuple[Any, ...]], merged_ranges: list[str]) -> int:
    if not values:
        return 1
    first_nonempty = next(
        (
            index + 1
            for index, row in enumerate(values[:10])
            if any(value is not None and str(value).strip() != "" for value in row)
        ),
        1,
    )
    for value in merged_ranges:
        _, min_row, _, max_row = range_boundaries(value)
        if min_row is not None and max_row is not None and min_row <= first_nonempty <= max_row:
            return first_nonempty
    candidates = [
        (index + 1, sum(value is not None and str(value).strip() != "" for value in row))
        for index, row in enumerate(values[:10])
    ]
    return max(candidates, key=lambda item: (item[1], -item[0]))[0]


def _xlsx_read(path: Path) -> TabularReadResult:
    workbook = load_workbook(path, read_only=False, data_only=True)
    all_rows: list[DatasetRow] = []
    sheets: list[SheetInspection] = []
    for worksheet in workbook.worksheets:
        values = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
        merged_ranges = [str(value) for value in worksheet.merged_cells.ranges]
        header_index = _header_row(values, merged_ranges)
        header_values = values[header_index - 1] if values else ()
        headers = _unique_headers(header_values)
        data_values = values[header_index:]
        rows = _dataset_rows(headers, data_values, start_row=header_index + 1)
        all_rows.extend(rows)
        sheets.append(
            SheetInspection(
                name=worksheet.title,
                header_row_index=header_index,
                headers=headers,
                row_count=len(data_values),
                column_count=len(headers),
                merged_ranges=merged_ranges,
                leading_rows=[list(row) for row in values[:5]],
            )
        )
    workbook.close()
    return TabularReadResult(_source_file(path), sheets, all_rows)


def read_tabular_file(path: str | Path) -> TabularReadResult:
    """Read a CSV or XLSX while preserving raw and normalized row representations."""

    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    return (
        _csv_read(source_path)
        if _format_for(source_path) == FileFormat.CSV
        else _xlsx_read(source_path)
    )

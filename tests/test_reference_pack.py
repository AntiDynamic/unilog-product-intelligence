"""Tests for Phase 6 Reference-Pack loading, normalization, and category-aware taxonomy."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from unilog_product_intelligence.application.product_truth import ProductTruthService
from unilog_product_intelligence.domain.models import Source, SourceAuthority, SourceType
from unilog_product_intelligence.domain.truth import ProductClassification, ProductTruth
from unilog_product_intelligence.enrichment.models import (
    Applicability,
    ReferenceAvailability,
)
from unilog_product_intelligence.enrichment.planner import AttributePlanner
from unilog_product_intelligence.enrichment.reference import (
    ReferencePack,
    ReferenceType,
    load_decimal_fraction,
    load_global_lov,
    load_manufacturer_brand,
    load_uom_master,
)


def _make_test_product(
    product_id: str = "prod-1",
    mpn: str = "ABC-123",
    desc: str = "Commercial brass pipe fitting",
    category: str = "Fittings",
    classpath: tuple[str, ...] = ("Plumbing", "Fittings"),
) -> ProductTruth:
    source = Source(
        source_id="mfg-src-1",
        source_type=SourceType.MANUFACTURER_PAGE,
        authority=SourceAuthority.HIGH,
        uri="https://manufacturer.example/product/1",
    )
    truth = ProductTruthService().create_from_raw_input(
        product_id,
        {"Mfg_Part_Num": mpn, "Part_Desc": desc, "Part_Manuf": "Mueller Industries"},
        source,
    )
    truth = ProductTruthService().add_classification(
        truth,
        ProductClassification(class_name=category, classpath=classpath),
    )
    return truth


# ==============================================================================
# TEST A — FILE DISCOVERY
# ==============================================================================


def test_reference_pack_discovery_and_classification(tmp_path: Path) -> None:
    """Discovers files and categorizes them under exact ReferenceTypes."""
    uom_file = tmp_path / "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx"
    dec_file = tmp_path / "Decimal_Fraction.xlsx"
    mfg_file = tmp_path / "UniCat_Manufacturer_and_Brand_List.xlsx"
    lov_file = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"

    # Create dummy workbooks
    for f in (uom_file, dec_file, mfg_file, lov_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["dummy"])
        wb.save(f)
        wb.close()

    pack = ReferencePack.discover([tmp_path])

    assert uom_file.name in pack.files
    assert dec_file.name in pack.files
    assert mfg_file.name in pack.files
    assert lov_file.name in pack.files

    # Unprovided files remain absent
    assert "FAUCETS_LOV.xlsx" not in pack.files
    assert "Fittings_LOV.xlsx" not in pack.files


# ==============================================================================
# TEST B — UOM LOADING
# ==============================================================================


def test_uom_loading_and_normalization(tmp_path: Path) -> None:
    """Loads approved UOM abbreviations, synonyms, measurement types, and capture forms."""
    uom_path = tmp_path / "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "UOM Master"
    ws.append(
        [
            "Approved UOM Abbreviation",
            "Measurement Type",
            "Capture Form",
            "Example",
            "Synonyms / Alternates",
            "Rule / Description",
        ]
    )
    ws.append(["IN", "Length", "in.", "5 IN, 1/2 IN", "inch, inches, IN.", "Standard inch"])
    ws.append(["FT", "Length", "ft.", "10 FT", "foot, feet", "Standard foot unit"])
    ws.append(
        ["PSI", "Pressure", "psi", "150 PSI", "lbs/sq in, pounds per square inch", "Pressure unit"]
    )
    ws.append(["GPM", "Flow Rate", "gpm", "2.5 GPM", "gallons per minute", "Gallons per min"])
    wb.save(uom_path)
    wb.close()

    uom_map = load_uom_master(uom_path)

    assert len(uom_map.records) == 4
    assert uom_map.is_approved("IN") is True
    assert uom_map.is_approved("in") is True
    assert uom_map.is_approved("UNKNOWN_UNIT") is False

    # Normalization of synonyms to canonical approved form
    assert uom_map.normalize("inches") == "IN"
    assert uom_map.normalize("inch") == "IN"
    assert uom_map.normalize("in.") == "IN"
    assert uom_map.normalize("IN") == "IN"
    assert uom_map.normalize("feet") == "FT"
    assert uom_map.normalize("gallons per minute") == "GPM"
    assert uom_map.normalize("unknown") is None

    # By measurement type
    length_uoms = uom_map.get_uoms_for_measurement("Length")
    assert "IN" in length_uoms
    assert "FT" in length_uoms


# ==============================================================================
# TEST C — UOM UNAVAILABLE
# ==============================================================================


def test_uom_unavailable_fails_closed(tmp_path: Path) -> None:
    """When UOM workbook is absent, loader fails closed without fabricating UOMs."""
    pack = ReferencePack.discover([tmp_path])

    assert pack.uom_available is False
    assert pack.status[ReferenceType.UOM_STANDARD] == ReferenceAvailability.REFERENCE_UNAVAILABLE
    assert pack.normalize_uom("inches") is None
    assert pack.is_approved_uom("IN") is False
    assert pack.get_allowed_uom("size") == ()


# ==============================================================================
# TEST D — FRACTION SIDE-BY-SIDE BLOCKS
# ==============================================================================


def test_decimal_fraction_side_by_side_blocks(tmp_path: Path) -> None:
    """Loads multiple side-by-side Fraction | Decimal column pairs."""
    dec_path = tmp_path / "Decimal_Fraction.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Decimal_Fraction"
    # Block 1 (Cols A-B), Blank (Col C), Block 2 (Cols D-E)
    ws.append(["Fraction", "Decimal", None, "Fraction", "Decimal"])
    ws.append(["1/2", 0.5, None, "1-1/2", 1.5])
    ws.append(["1/4", 0.25, None, "2-1/4", 2.25])
    ws.append(["3/4", 0.75, None, "50-1/4", 50.25])
    ws.append(["5/16", 0.3125, None, "1/8", 0.125])
    wb.save(dec_path)
    wb.close()

    frac_map = load_decimal_fraction(dec_path)

    # Block 1 values
    assert frac_map.to_decimal("1/2") == 0.5
    assert frac_map.to_decimal("1/4") == 0.25
    assert frac_map.to_decimal("5/16") == 0.3125
    assert frac_map.to_fraction(0.5) == "1/2"
    assert frac_map.to_fraction("0.25") == "1/4"

    # Block 2 values
    assert frac_map.to_decimal("1-1/2") == 1.5
    assert frac_map.to_decimal("1 1/2") == 1.5
    assert frac_map.to_decimal("50-1/4") == 50.25
    assert frac_map.to_fraction(50.25) == "50-1/4"
    assert frac_map.to_fraction(1.5) in {"1-1/2", "1 1/2"}


# ==============================================================================
# TEST E — MANUFACTURER / BRAND MASTER
# ==============================================================================


def test_manufacturer_brand_master_loading(tmp_path: Path) -> None:
    """Loads exact canonical manufacturer and brand names with case-insensitive resolution."""
    mfg_path = tmp_path / "UniCat_Manufacturer_and_Brand_List.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Mfg_Brand_List"
    ws.append(["MANUFACTURER_NAME", "MANUFACTURER_CODE", "BRAND_NAME", "BRAND_CODE"])
    ws.append(["3M Company", "3M01", "3M™", "BR3M"])
    ws.append(["Freud America, Inc.", "FR01", "Diablo", "BRDIA"])
    ws.append(["Milwaukee Electric Tool Corp.", "MIL01", "Milwaukee", "BRMIL"])
    ws.append(["Jam Industrial Supply LLC", "JIS01", "Jam Supply", "BRJAM"])
    wb.save(mfg_path)
    wb.close()

    mfg_index = load_manufacturer_brand(mfg_path)

    assert len(mfg_index.records) == 4

    # Exact canonical preservation
    assert mfg_index.resolve_manufacturer("3M Company") == "3M Company"
    assert mfg_index.resolve_manufacturer("3m company") == "3M Company"
    assert mfg_index.resolve_manufacturer("3M01") == "3M Company"
    assert mfg_index.resolve_manufacturer("Freud America") == "Freud America, Inc."

    assert mfg_index.resolve_brand("diablo") == "Diablo"
    assert mfg_index.resolve_brand("milwaukee") == "Milwaukee"
    assert mfg_index.resolve_brand("BR3M") == "3M™"

    # Pair resolution
    mfg, brand = mfg_index.resolve_pair("Diablo")
    assert brand == "Diablo"
    assert mfg == "Freud America, Inc."


# ==============================================================================
# TEST F — GLOBAL LOV TAXONOMY CONTEXT SEPARATION
# ==============================================================================


def test_global_lov_taxonomy_context_separation(tmp_path: Path) -> None:
    """Distinct categories sharing the same attribute name retain independent allowed values."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Global LOV"
    ws.append(
        [
            "Classpath",
            "Leaf Node",
            "Attribute Label",
            "Attribute Values",
            "Filtering Y/N",
            "Guidelines",
            "Remarks",
            "UOM",
        ]
    )
    # Category 1: Plumbing > Fittings
    ws.append(
        [
            "Plumbing > Fittings",
            "Fittings",
            "Material",
            "Brass\nBronze\nCopper\nPVC\nStainless Steel",
            "Y",
            "Select fitting body material",
            None,
            None,
        ]
    )
    ws.append(
        [
            "Plumbing > Fittings",
            "Fittings",
            "Size",
            "1/2\n3/4\n1\n2",
            "Y",
            "Nominal pipe size",
            None,
            "IN",
        ]
    )

    # Category 2: Fasteners > Bolts
    ws.append(
        [
            "Fasteners > Bolts",
            "Hex Bolts",
            "Material",
            "Grade 5 Steel\nGrade 8 Steel\nTitanium\nZinc Plated",
            "Y",
            "Bolt material grade",
            None,
            None,
        ]
    )
    ws.append(
        [
            "Fasteners > Bolts",
            "Hex Bolts",
            "Size",
            "1/4-20\n3/8-16\n1/2-13",
            "Y",
            "Thread size",
            None,
            "IN",
        ]
    )

    wb.save(lov_path)
    wb.close()

    lov_index = load_global_lov(lov_path)

    # Category 1 values for Material
    fittings_materials = lov_index.get_allowed_values(
        "Material", classpath=("Plumbing", "Fittings")
    )
    assert set(fittings_materials) == {"Brass", "Bronze", "Copper", "PVC", "Stainless Steel"}
    assert "Titanium" not in fittings_materials

    # Category 2 values for Material
    bolt_materials = lov_index.get_allowed_values(
        "Material", classpath=("Fasteners", "Hex Bolts"), category="Hex Bolts"
    )
    assert set(bolt_materials) == {"Grade 5 Steel", "Grade 8 Steel", "Titanium", "Zinc Plated"}
    assert "Brass" not in bolt_materials

    # UOM lookup
    assert lov_index.get_allowed_uom("Size", classpath=("Plumbing", "Fittings")) == ("IN",)


# ==============================================================================
# TEST G — CATEGORY LOV PACKS
# ==============================================================================


def test_category_lov_packs_loading(tmp_path: Path) -> None:
    """Loads separate category-specific reference workbooks."""
    faucets_path = tmp_path / "FAUCETS_LOV.xlsx"
    fittings_path = tmp_path / "Fittings_LOV.xlsx"

    # Create Faucets LOV
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "LOV Detail"
    ws1.append(["Attribute Label", "Attribute Values", "Filtering Y/N", "UOM"])
    ws1.append(["Faucet Type", "Bridge\nCenterset\nSingle Hole\nWall Mount", "Y", None])
    ws1.append(["Finish", "Chrome\nMatte Black\nBrushed Nickel\nPolished Brass", "Y", None])
    wb1.save(faucets_path)
    wb1.close()

    # Create Fittings LOV
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Attribute Detail"
    ws2.append(["Attribute Label", "Attribute Values", "Filtering Y/N", "UOM"])
    ws2.append(["Fitting Type", "Adapter\nCoupling\nElbow\nTee\nUnion", "Y", None])
    ws2.append(["Connection Type", "NPT x NPT\nPush-Fit\nSweat x Sweat", "Y", None])
    wb2.save(fittings_path)
    wb2.close()

    pack = ReferencePack.discover([tmp_path])

    assert "faucets_lov.xlsx" in [f.casefold() for f in pack.files]
    assert "fittings_lov.xlsx" in [f.casefold() for f in pack.files]

    # Category-specific queries
    faucet_types = pack.get_allowed_values("Faucet Type", category="Faucets")
    assert "Centerset" in faucet_types
    assert "Wall Mount" in faucet_types

    fitting_types = pack.get_allowed_values("Fitting Type", category="Fittings")
    assert "Coupling" in fitting_types
    assert "Elbow" in fitting_types


# ==============================================================================
# TEST H — MULTI-ROW / MERGED HEADERS
# ==============================================================================


def test_multi_row_and_merged_header_handling(tmp_path: Path) -> None:
    """Handles merged cells and multi-row layout where each value is on a separate row."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: Superheader title
    ws.append(["UNILOG TAXONOMY MASTER & CONTROLLED VOCABULARY", None, None, None])
    # Row 2: Actual column headers
    ws.append(["Classpath", "Leaf Node", "Attribute Label", "Attribute Values", "UOM"])
    # Rows 3-7: Material with values spread across multiple rows (merged-style empty cells)
    ws.append(["Plumbing > Valves", "Ball Valves", "Material", "Brass", None])
    ws.append([None, None, None, "Bronze", None])
    ws.append([None, None, None, "Cast Iron", None])
    ws.append([None, None, None, "Stainless Steel", None])
    # Rows 8-10: Valve Size
    ws.append([None, None, "Size", "1/2", "IN"])
    ws.append([None, None, None, "3/4", "IN"])
    ws.append([None, None, None, "1", "IN"])

    wb.save(lov_path)
    wb.close()

    lov_index = load_global_lov(lov_path)

    valve_materials = lov_index.get_allowed_values("Material", classpath=("Plumbing", "Valves"))
    assert set(valve_materials) == {"Brass", "Bronze", "Cast Iron", "Stainless Steel"}

    valve_sizes = lov_index.get_allowed_values("Size", classpath=("Plumbing", "Valves"))
    assert set(valve_sizes) == {"1/2", "3/4", "1"}
    assert lov_index.get_allowed_uom("Size", classpath=("Plumbing", "Valves")) == ("IN",)


# ==============================================================================
# TEST I — PER-REFERENCE AVAILABILITY TRACKING
# ==============================================================================


def test_per_reference_availability_fail_closed(tmp_path: Path) -> None:
    """Each reference type is tracked independently; missing files remain UNAVAILABLE."""
    # Only supply UOM workbook
    uom_path = tmp_path / "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Approved UOM Abbreviation", "Measurement Type"])
    ws.append(["IN", "Length"])
    wb.save(uom_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])

    assert pack.uom_available is True
    assert pack.status[ReferenceType.UOM_STANDARD] == ReferenceAvailability.REFERENCE_AVAILABLE

    assert pack.lov_available is False
    assert pack.status[ReferenceType.GLOBAL_LOV] == (ReferenceAvailability.REFERENCE_UNAVAILABLE)
    assert pack.status[ReferenceType.DECIMAL_FRACTION] == (
        ReferenceAvailability.REFERENCE_UNAVAILABLE
    )
    assert pack.status[ReferenceType.MANUFACTURER_BRAND] == (
        ReferenceAvailability.REFERENCE_UNAVAILABLE
    )

    # Overall is available since 1 file was found, but individual status is accurate
    assert pack.available is True


# ==============================================================================
# TEST J — ATTRIBUTE PLANNER CONSUMPTION
# ==============================================================================


def test_attribute_planner_consumes_category_reference(tmp_path: Path) -> None:
    """AttributePlanner populates allowed_values and allowed_uom from category reference pack."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    uom_path = tmp_path / "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx"

    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(["Classpath", "Leaf Node", "Attribute Label", "Attribute Values", "UOM"])
    ws1.append(
        [
            "Plumbing > Fittings",
            "Fittings",
            "Fitting Type",
            "Adapter\nCoupling\nElbow\nTee",
            None,
        ]
    )
    ws1.append(
        [
            "Plumbing > Fittings",
            "Fittings",
            "Connection Type",
            "NPT x NPT\nPush-Fit\nSweat",
            None,
        ]
    )
    ws1.append(
        [
            "Plumbing > Fittings",
            "Fittings",
            "Material",
            "Brass\nBronze\nCopper\nPVC",
            None,
        ]
    )
    ws1.append(["Plumbing > Fittings", "Fittings", "Size", "1/2\n3/4\n1\n2", "IN"])
    wb1.save(lov_path)
    wb1.close()

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Approved UOM Abbreviation", "Measurement Type"])
    ws2.append(["IN", "Length"])
    wb2.save(uom_path)
    wb2.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)

    product = _make_test_product(category="Fittings", classpath=("Plumbing", "Fittings"))
    plans = planner.plan(product)

    plan_by_attr = {p.attribute_name: p for p in plans}

    # Material plan
    mat_plan = plan_by_attr["Material"]
    assert set(mat_plan.allowed_values) == {"Brass", "Bronze", "Copper", "PVC"}
    assert mat_plan.reference_availability == ReferenceAvailability.REFERENCE_AVAILABLE
    assert "lov" in mat_plan.validation_requirements

    # Size plan
    size_plan = plan_by_attr["Size"]
    assert set(size_plan.allowed_values) == {"1/2", "3/4", "1", "2"}
    assert size_plan.allowed_uom == ("IN",)
    assert size_plan.reference_availability == ReferenceAvailability.REFERENCE_AVAILABLE

    # Fitting Type plan
    fit_plan = plan_by_attr["Fitting Type"]
    assert set(fit_plan.allowed_values) == {"Adapter", "Coupling", "Elbow", "Tee"}
    assert fit_plan.applicability == Applicability.REQUIRED

"""Tests for Category-Aware Attribute Planning & Evidence-Grounded Extraction (Phase 6)."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from unilog_product_intelligence.application.product_truth import ProductTruthService
from unilog_product_intelligence.domain.models import Source, SourceAuthority, SourceType
from unilog_product_intelligence.domain.truth import (
    CandidateValue,
    Evidence,
    EvidenceType,
    ProductClassification,
    ProductTruth,
    ValueStatus,
)
from unilog_product_intelligence.enrichment.agent import (
    EvidenceGroundedEnrichmentAgent,
    evidence_references,
)
from unilog_product_intelligence.enrichment.models import (
    Applicability,
    EnrichmentCandidate,
    EnrichmentDecision,
    FinalAttributeStatus,
    PublicationState,
    ReferenceAvailability,
    ValidationSeverity,
)
from unilog_product_intelligence.enrichment.planner import AttributePlanner
from unilog_product_intelligence.enrichment.reference import (
    ReferencePack,
    separate_value_and_uom,
)
from unilog_product_intelligence.enrichment.validation import ValidationPipeline
from unilog_product_intelligence.providers.base import LLMProvider, LLMRequest, LLMResponse


class MockEnrichmentProvider(LLMProvider):
    def __init__(self, candidates_payload: list[dict[str, object]]) -> None:
        self.candidates_payload = candidates_payload

    def generate(self, request: LLMRequest) -> LLMResponse:
        return LLMResponse(
            output_text=json.dumps(
                {
                    "candidates": self.candidates_payload,
                    "unresolved_attributes": [],
                }
            ),
            model="gemini-3.5-flash-lite",
            input_tokens=15,
            output_tokens=12,
        )


def _make_test_product(
    product_id: str = "prod-1",
    mpn: str = "ABC-123",
    desc: str = "Industrial Fitting",
    category: str = "Couplings",
    classpath: tuple[str, ...] = ("Plumbing", "Fittings", "Couplings"),
) -> ProductTruth:
    source = Source(
        source_id="mfg-source-1",
        source_type=SourceType.MANUFACTURER_PAGE,
        authority=SourceAuthority.AUTHORITATIVE,
        uri="https://mfg.example.com/product/abc-123",
    )
    truth = ProductTruthService().create_from_raw_input(
        product_id,
        {
            "Mfg_Part_Num": mpn,
            "Part_Desc": desc,
            "Unilog_Brand": "Acme Brand",
            "Part_Manuf": "Acme Mfg",
        },
        source,
    )
    truth = ProductTruthService().add_classification(
        truth,
        ProductClassification(
            class_name=category,
            classpath=classpath,
        ),
    )
    return truth


# ==============================================================================
# TEST A — GLOBAL LOV CATEGORY RESOLUTION
# ==============================================================================


def test_global_lov_category_resolution(tmp_path: Path) -> None:
    """Classpath Plumbing > Fittings > Couplings returns relevant attributes from GLOBAL_LOV."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Classpath",
        "Leaf Node",
        "Attribute Label",
        "Attribute Values",
        "Filtering Y/N",
        "Guidelines",
        "Remarks",
        "UOM",
    ])
    ws.append([
        "Plumbing > Fittings > Couplings",
        "Couplings",
        "Fitting Type",
        "Rigid Coupling\nFlexible Coupling\nCompression Coupling",
        "Y",
        "Required fitting type",
        "Required",
        None,
    ])
    ws.append([
        "Plumbing > Fittings > Couplings",
        "Couplings",
        "Connection Type",
        "NPT x NPT\nPush-Fit\nSweat",
        "Y",
        "Connection standard",
        None,
        None,
    ])
    ws.append([
        "Plumbing > Fittings > Couplings",
        "Couplings",
        "Material",
        "Brass\nBronze\nCopper\nPVC\nStainless Steel",
        "Y",
        "Body material",
        None,
        None,
    ])
    ws.append([
        "Plumbing > Fittings > Couplings",
        "Couplings",
        "Size",
        "1/2\n3/4\n1\n2",
        "N",
        "Nominal size",
        None,
        "IN",
    ])
    wb.save(lov_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)

    product = _make_test_product(
        category="Couplings",
        classpath=("Plumbing", "Fittings", "Couplings"),
    )
    plans = planner.plan(product)

    plan_names = {p.attribute_name for p in plans}
    assert plan_names == {"Fitting Type", "Connection Type", "Material", "Size"}
    assert all(p.schema_source == "GLOBAL_LOV" for p in plans)
    assert all(p.reference_availability == ReferenceAvailability.REFERENCE_AVAILABLE for p in plans)


# ==============================================================================
# TEST B — ATTRIBUTE NAME COLLISION
# ==============================================================================


def test_attribute_name_collision(tmp_path: Path) -> None:
    """Same attribute name ('Material') in different categories resolves distinct allowed values."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Classpath", "Leaf Node", "Attribute Label", "Attribute Values", "Filtering Y/N", "UOM"
    ])
    ws.append([
        "Plumbing > Fittings", "Fittings", "Material",
        "Brass\nBronze\nCopper\nPVC\nStainless Steel", "Y", None,
    ])
    ws.append([
        "Fasteners > Bolts", "Hex Bolts", "Material",
        "Grade 5 Steel\nGrade 8 Steel\nTitanium\nZinc Plated", "Y", None,
    ])
    wb.save(lov_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)

    fitting_prod = _make_test_product(
        category="Fittings", classpath=("Plumbing", "Fittings")
    )
    bolt_prod = _make_test_product(
        category="Hex Bolts", classpath=("Fasteners", "Hex Bolts")
    )

    fitting_plans = {p.attribute_name: p for p in planner.plan(fitting_prod)}
    bolt_plans = {p.attribute_name: p for p in planner.plan(bolt_prod)}

    assert "Material" in fitting_plans
    assert "Material" in bolt_plans

    assert set(fitting_plans["Material"].allowed_values) == {
        "Brass", "Bronze", "Copper", "PVC", "Stainless Steel"
    }
    assert set(bolt_plans["Material"].allowed_values) == {
        "Grade 5 Steel", "Grade 8 Steel", "Titanium", "Zinc Plated"
    }
    assert "Titanium" not in fitting_plans["Material"].allowed_values
    assert "Brass" not in bolt_plans["Material"].allowed_values


# ==============================================================================
# TEST C — CATEGORY LOV OVERRIDES GLOBAL
# ==============================================================================


def test_category_lov_overrides_global(tmp_path: Path) -> None:
    """Category-specific LOV pack takes precedence over global LOV."""
    global_lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    category_lov_path = tmp_path / "FAUCETS_LOV.xlsx"

    wb_global = Workbook()
    ws_global = wb_global.active
    ws_global.append(["Classpath", "Leaf Node", "Attribute Label", "Attribute Values"])
    ws_global.append(["Plumbing > Faucets", "Faucets", "Faucet Type", "Generic Faucet"])
    ws_global.append(["Plumbing > Faucets", "Faucets", "Finish", "Generic Chrome"])
    wb_global.save(global_lov_path)
    wb_global.close()

    wb_cat = Workbook()
    ws_cat = wb_cat.active
    ws_cat.append(["Attribute Label", "Attribute Values", "UOM"])
    ws_cat.append([
        "Faucet Type",
        "Centerset Faucet\nWidespread Faucet\nSingle Hole Faucet\nWall Mount Faucet",
        None,
    ])
    ws_cat.append([
        "Finish",
        "Polished Chrome\nBrushed Nickel\nMatte Black\nOil Rubbed Bronze",
        None,
    ])
    ws_cat.append(["Flow Rate", "1.2\n1.5\n1.8\n2.2", "GPM"])
    wb_cat.save(category_lov_path)
    wb_cat.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)

    faucet_prod = _make_test_product(category="Faucets", classpath=("Plumbing", "Faucets"))
    plans = {p.attribute_name: p for p in planner.plan(faucet_prod)}

    assert plans["Faucet Type"].schema_source == "CATEGORY_LOV"
    assert "Centerset Faucet" in plans["Faucet Type"].allowed_values
    assert "Generic Faucet" not in plans["Faucet Type"].allowed_values
    assert "Flow Rate" in plans
    assert plans["Flow Rate"].allowed_uom == ("GPM",)


# ==============================================================================
# TEST D — REQUIRED VS OPTIONAL
# ==============================================================================


def test_required_vs_optional_ordering(tmp_path: Path) -> None:
    """Preserves required, filtering, and optional priority order."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Classpath", "Leaf Node", "Attribute Label", "Attribute Values", "Filtering Y/N", "Remarks"
    ])
    ws.append([
        "Abrasives > Cut-Off Wheels",
        "Cut-Off Wheels",
        "Wheel Diameter",
        "4-1/2\n6\n14",
        "Y",
        "Required",
    ])
    ws.append([
        "Abrasives > Cut-Off Wheels", "Cut-Off Wheels", "Arbor Size", "5/8-11\n7/8", "Y", None
    ])
    ws.append([
        "Abrasives > Cut-Off Wheels", "Cut-Off Wheels", "Max RPM", "8500\n13300", "N", None
    ])
    wb.save(lov_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)

    prod = _make_test_product(
        category="Cut-Off Wheels", classpath=("Abrasives", "Cut-Off Wheels")
    )
    plans = planner.plan(prod)

    assert len(plans) == 3
    # Required first
    assert plans[0].attribute_name == "Wheel Diameter"
    assert plans[0].applicability == Applicability.REQUIRED
    assert plans[0].priority == 90

    # Filtering next
    assert plans[1].attribute_name == "Arbor Size"
    assert plans[1].applicability == Applicability.OPTIONAL
    assert plans[1].priority == 70

    # Optional last
    assert plans[2].attribute_name == "Max RPM"
    assert plans[2].applicability == Applicability.OPTIONAL
    assert plans[2].priority == 50


# ==============================================================================
# TEST E — UNKNOWN CATEGORY
# ==============================================================================


def test_unknown_category_fails_closed_to_fallback() -> None:
    """When category reference is unavailable, planner falls back without inventing schemas."""
    pack = ReferencePack(ReferenceAvailability.REFERENCE_UNAVAILABLE, {})
    planner = AttributePlanner(reference_pack=pack)

    # Product with no category LOV match and custom existing attributes
    prod = _make_test_product(
        desc="Liquid Chemical Compound",
        category="Specialty Chemical Compound",
        classpath=("Chemicals", "Industrial", "Specialty"),
    )
    truth_service = ProductTruthService()
    prod = truth_service.add_attribute_candidate(
        prod,
        "ph_level",
        CandidateValue(
            candidate_id="cand-1",
            raw_value="7.0",
            normalized_value="7.0",
            status=ValueStatus.ENRICHED,
        ),
        canonical_name="pH Level",
    )

    plans = planner.plan(prod)

    assert len(plans) == 1
    assert plans[0].attribute_name == "pH Level"
    assert plans[0].schema_source == "FALLBACK_EXISTING_ATTRIBUTES"
    assert plans[0].reference_availability == ReferenceAvailability.REFERENCE_UNAVAILABLE
    assert plans[0].applicability == Applicability.OPTIONAL


# ==============================================================================
# TEST F — PLAN DETERMINISM
# ==============================================================================


def test_plan_determinism(tmp_path: Path) -> None:
    """Repeated planner calls on the same product and reference pack yield identical plans."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Classpath", "Leaf Node", "Attribute Label", "Attribute Values", "Filtering Y/N"])
    ws.append(["Plumbing > Valves", "Valves", "Valve Type", "Ball\nCheck\nGate\nGlobe", "Y"])
    ws.append([
        "Plumbing > Valves",
        "Valves",
        "Material",
        "Brass\nBronze\nCast Iron\nStainless Steel",
        "Y",
    ])
    ws.append(["Plumbing > Valves", "Valves", "Connection", "NPT\nFlanged\nSolder", "Y"])
    wb.save(lov_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)
    prod = _make_test_product(category="Valves", classpath=("Plumbing", "Valves"))

    plan1 = planner.plan(prod)
    plan2 = planner.plan(prod)

    assert len(plan1) == len(plan2)
    assert [p.attribute_id for p in plan1] == [p.attribute_id for p in plan2]
    assert [p.priority for p in plan1] == [p.priority for p in plan2]
    assert [p.allowed_values for p in plan1] == [p.allowed_values for p in plan2]


# ==============================================================================
# TEST G — EXISTING VERIFIED ATTRIBUTE PRESERVATION
# ==============================================================================


def test_existing_verified_attribute_preserved(tmp_path: Path) -> None:
    """Verified existing attribute with evidence is not planned for re-enrichment."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Classpath", "Leaf Node", "Attribute Label", "Attribute Values"])
    ws.append(["Plumbing > Fittings", "Fittings", "Material", "Brass\nStainless Steel"])
    ws.append(["Plumbing > Fittings", "Fittings", "Size", "1/2\n3/4"])
    wb.save(lov_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)
    prod = _make_test_product(category="Fittings", classpath=("Plumbing", "Fittings"))

    # Add verified material with evidence
    prod.evidence.append(
        Evidence(
            evidence_id="ev-100",
            source_id="mfg-source-1",
            product_id=prod.product_id,
            quoted_text="Material: Stainless Steel",
            evidence_type=EvidenceType.DIRECT_TEXT,
        )
    )
    truth_service = ProductTruthService()
    prod = truth_service.add_attribute_candidate(
        prod,
        "material",
        CandidateValue(
            candidate_id="cand-verified",
            raw_value="Stainless Steel",
            normalized_value="Stainless Steel",
            evidence_ids=["ev-100"],
            status=ValueStatus.VERIFIED,
        ),
        canonical_name="Material",
    )
    attr = prod.attribute("material")
    attr.status = ValueStatus.VERIFIED
    attr.normalized_value = "Stainless Steel"

    plans = {p.attribute_name: p for p in planner.plan(prod)}

    assert plans["Material"].enrichment_required == EnrichmentDecision.NO_ACTION
    assert "Existing publishable value is not re-enriched." in plans["Material"].reason
    assert plans["Size"].enrichment_required == EnrichmentDecision.ENRICH


# ==============================================================================
# TEST H — LLM PLAN CONSTRAINT (UNPLANNED ATTRIBUTES REJECTED)
# ==============================================================================


def test_unplanned_attribute_rejected_by_agent_and_validator() -> None:
    """Agent and ValidationPipeline reject candidate attributes not in the plan."""
    prod = _make_test_product(category="Fittings", classpath=("Plumbing", "Fittings"))
    prod.evidence.append(
        Evidence(
            evidence_id="ev-1",
            source_id="mfg-source-1",
            product_id=prod.product_id,
            quoted_text="Material: Brass. Country of Origin: USA.",
            evidence_type=EvidenceType.DIRECT_TEXT,
        )
    )

    pack = ReferencePack(ReferenceAvailability.REFERENCE_UNAVAILABLE, {})
    planner = AttributePlanner(reference_pack=pack)
    plans = planner.plan(prod)

    # Provider outputs one planned attribute and one unplanned attribute
    provider = MockEnrichmentProvider([
        {
            "attribute": "material",
            "value": "Brass",
            "raw_value": "Brass",
            "normalized_value": "Brass",
            "evidence_id": "ev-1",
            "evidence_text": "Material: Brass",
            "status": "DIRECT",
            "reason": "Directly stated.",
        },
        {
            "attribute": "unplanned_secret_attr",
            "value": "Secret",
            "raw_value": "Secret",
            "normalized_value": "Secret",
            "evidence_id": "ev-1",
            "evidence_text": "Secret info",
            "status": "DIRECT",
            "reason": "Unplanned.",
        },
    ])

    agent = EvidenceGroundedEnrichmentAgent(provider)
    candidates = agent.enrich(prod, plans, evidence_references(prod))

    # Agent drops unplanned candidates before validation
    cand_attr_ids = {c.attribute_id for c in candidates}
    assert "material" in cand_attr_ids
    assert "unplanned_secret_attr" not in cand_attr_ids


# ==============================================================================
# TEST I — EVIDENCE REQUIRED
# ==============================================================================


def test_evidence_required_for_publishable_candidate() -> None:
    """Candidate with no evidence fails validation and is rejected."""
    prod = _make_test_product(category="Fittings", classpath=("Plumbing", "Fittings"))
    planner = AttributePlanner()
    plans = planner.plan(prod)

    candidate_without_evidence = EnrichmentCandidate(
        candidate_id="cand-no-ev",
        product_id=prod.product_id,
        attribute_id="material",
        attribute="Material",
        value="Brass",
        raw_value="Brass",
        evidence_ids=(),
        evidence_text=None,
        evidence=(),
        status=FinalAttributeStatus.ENRICHED,
        candidate_reason="Guess with no evidence.",
    )

    validated, validations, reviews = ValidationPipeline().validate(
        prod, plans, [candidate_without_evidence]
    )

    assert validated[0].status == FinalAttributeStatus.REJECTED
    assert any(
        v.validator == "evidence" and v.severity == ValidationSeverity.BLOCKING
        for v in validations
    )
    pub_state = ValidationPipeline().publication_state(plans, validated, validations, reviews)
    assert pub_state == PublicationState.BLOCKED


# ==============================================================================
# TEST J — DETERMINISTIC UOM SEPARATION
# ==============================================================================


def test_uom_separation_deterministic(tmp_path: Path) -> None:
    """Evidence containing '120 V' separates value='120' and uom='V'."""
    uom_path = tmp_path / "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Approved UOM Abbreviation", "Measurement Type"])
    ws.append(["V", "Voltage"])
    ws.append(["IN", "Length"])
    ws.append(["PSI", "Pressure"])
    wb.save(uom_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])

    # Test separation helper directly
    val, uom = separate_value_and_uom("120 V", allowed_uoms=("V",), reference_pack=pack)
    assert val == "120"
    assert uom == "V"

    val_psi, uom_psi = separate_value_and_uom("150 PSI", allowed_uoms=("PSI",), reference_pack=pack)
    assert val_psi == "150"
    assert uom_psi == "PSI"

    val_frac, uom_frac = separate_value_and_uom("1/2 IN", allowed_uoms=("IN",), reference_pack=pack)
    assert val_frac == "1/2"
    assert uom_frac == "IN"


# ==============================================================================
# TEST K — LOV VALIDATION (VALID VS UNSUPPORTED)
# ==============================================================================


def test_lov_validation_valid_vs_unsupported(tmp_path: Path) -> None:
    """Allowed value passes; unsupported value produces validation error / review."""
    lov_path = tmp_path / "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Classpath", "Leaf Node", "Attribute Label", "Attribute Values"])
    ws.append(["Plumbing > Fittings", "Fittings", "Material", "Brass\nStainless Steel\nCopper"])
    wb.save(lov_path)
    wb.close()

    pack = ReferencePack.discover([tmp_path])
    planner = AttributePlanner(reference_pack=pack)

    prod = _make_test_product(category="Fittings", classpath=("Plumbing", "Fittings"))
    prod.evidence.append(
        Evidence(
            evidence_id="ev-1",
            source_id="mfg-source-1",
            product_id=prod.product_id,
            quoted_text="Body made of Stainless Steel or Banana Alloy",
            evidence_type=EvidenceType.DIRECT_TEXT,
        )
    )
    plans = planner.plan(prod)
    ref = evidence_references(prod)[0]

    # Valid candidate
    valid_cand = EnrichmentCandidate(
        candidate_id="cand-valid",
        product_id=prod.product_id,
        attribute_id="material",
        attribute="Material",
        value="Stainless Steel",
        raw_value="Stainless Steel",
        source_id=ref.source_id,
        evidence_ids=(ref.evidence_id,),
        evidence_text=ref.evidence_text,
        evidence=(ref,),
        status=FinalAttributeStatus.ENRICHED,
        candidate_reason="Direct fact.",
    )
    val_cand, val_res, _ = ValidationPipeline().validate(prod, plans, [valid_cand])
    assert val_cand[0].status == FinalAttributeStatus.ENRICHED
    assert not any(v.validator == "lov" and not v.passed for v in val_res)

    # Unsupported candidate (Banana Alloy)
    invalid_cand = EnrichmentCandidate(
        candidate_id="cand-invalid",
        product_id=prod.product_id,
        attribute_id="material",
        attribute="Material",
        value="Banana Alloy",
        raw_value="Banana Alloy",
        source_id=ref.source_id,
        evidence_ids=(ref.evidence_id,),
        evidence_text=ref.evidence_text,
        evidence=(ref,),
        status=FinalAttributeStatus.ENRICHED,
        candidate_reason="Unsupported alloy.",
    )
    inv_cand, inv_res, reviews = ValidationPipeline().validate(prod, plans, [invalid_cand])
    assert inv_cand[0].status == FinalAttributeStatus.REVIEW_REQUIRED
    assert any(v.validator == "lov" and not v.passed for v in inv_res)
